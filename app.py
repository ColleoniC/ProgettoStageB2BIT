import io
import os
import secrets
import sqlite3
import smtplib
from email.message import EmailMessage
from datetime import date, datetime, timedelta

from flask import Flask, Response, flash, jsonify, redirect, render_template, request, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash, generate_password_hash

from dotenv import load_dotenv
load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', '')

ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', '')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', '')

EMAIL_MITTENTE = os.environ.get('EMAIL_MITTENTE', '')
EMAIL_APP_PASSWORD = os.environ.get('EMAIL_APP_PASSWORD', '')

def get_db():
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    return conn


def invia_email(destinatario, oggetto, corpo):
    """Compone e invia un'email tramite Gmail SMTP.
    Usata sia per la password provvisoria che per le segnalazioni."""
    msg = EmailMessage()
    msg['Subject'] = oggetto
    msg['From'] = EMAIL_MITTENTE
    msg['To'] = destinatario
    msg.set_content(corpo)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_MITTENTE, EMAIL_APP_PASSWORD)
        smtp.send_message(msg)


def init_db():
    """Crea le tabelle (se non esistono) e aggiunge le colonne mancanti."""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            name TEXT
        )
    ''')

    cursor.execute('PRAGMA table_info(users)')
    colonne_esistenti = {riga[1] for riga in cursor.fetchall()}
    colonne_richieste = {
        'cognome': 'TEXT',
        'orario_inizio': 'TEXT',
        'orario_pausa_inizio': 'TEXT',
        'orario_pausa_fine': 'TEXT',
        'orario_fine': 'TEXT',
        'privacy_accettata': 'INTEGER DEFAULT 0',
    }
    for nome_colonna, tipo in colonne_richieste.items():
        if nome_colonna not in colonne_esistenti:
            cursor.execute(f'ALTER TABLE users ADD COLUMN {nome_colonna} {tipo}')
            print(f"Colonna aggiunta: {nome_colonna}")

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS timbrature (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            creato_il TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')

    cursor.execute('PRAGMA table_info(timbrature)')
    colonne_timbrature = {riga[1] for riga in cursor.fetchall()}
    for nome_colonna, tipo_colonna in {'lat': 'REAL', 'lng': 'REAL'}.items():
        if nome_colonna not in colonne_timbrature:
            cursor.execute(f'ALTER TABLE timbrature ADD COLUMN {nome_colonna} {tipo_colonna}')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS assenze_segnalate (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            data_riferimento TEXT NOT NULL,
            creato_il TEXT NOT NULL,
            UNIQUE(user_id, data_riferimento)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS notifiche (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            messaggio TEXT NOT NULL,
            tipo TEXT NOT NULL DEFAULT 'info',
            creato_il TEXT NOT NULL,
            letta INTEGER NOT NULL DEFAULT 0,
            user_id INTEGER,
            data_riferimento TEXT
        )
    ''')

    cursor.execute('PRAGMA table_info(notifiche)')
    colonne_notifiche = {riga[1] for riga in cursor.fetchall()}
    for nome_colonna, tipo_colonna in {'user_id': 'INTEGER', 'data_riferimento': 'TEXT', 'segnalata': 'INTEGER DEFAULT 0', 'lat': 'REAL', 'lng': 'REAL'}.items():
        if nome_colonna not in colonne_notifiche:
            cursor.execute(f'ALTER TABLE notifiche ADD COLUMN {nome_colonna} {tipo_colonna}')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS messaggi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            mittente TEXT NOT NULL,
            testo TEXT NOT NULL,
            creato_il TEXT NOT NULL,
            letto_admin INTEGER NOT NULL DEFAULT 0,
            letto_utente INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS orari_giornalieri (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            orario_inizio TEXT,
            orario_pausa_inizio TEXT,
            orario_pausa_fine TEXT,
            orario_fine TEXT,
            note TEXT,
            creato_il TEXT NOT NULL,
            UNIQUE(user_id, data),
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')

    cursor.execute('PRAGMA table_info(orari_giornalieri)')
    colonne_orari_giornalieri = {riga[1] for riga in cursor.fetchall()}
    for nome_colonna, tipo_colonna in {'permesso_inizio': 'TEXT', 'permesso_fine': 'TEXT'}.items():
        if nome_colonna not in colonne_orari_giornalieri:
            cursor.execute(f'ALTER TABLE orari_giornalieri ADD COLUMN {nome_colonna} {tipo_colonna}')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS richieste (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            data_inizio TEXT NOT NULL,
            data_fine TEXT NOT NULL,
            motivo TEXT,
            stato TEXT NOT NULL DEFAULT 'in_attesa',
            creato_il TEXT NOT NULL,
            gestita_il TEXT,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')

    cursor.execute('PRAGMA table_info(richieste)')
    colonne_richieste_tab = {riga[1] for riga in cursor.fetchall()}
    for nome_colonna, tipo_colonna in {
        'ora_inizio': 'TEXT',
        'ora_fine': 'TEXT',
        'rimozione_richiesta': 'INTEGER NOT NULL DEFAULT 0',
    }.items():
        if nome_colonna not in colonne_richieste_tab:
            cursor.execute(f'ALTER TABLE richieste ADD COLUMN {nome_colonna} {tipo_colonna}')

    conn.commit()
    conn.close()


def crea_notifica(messaggio, tipo='info', user_id=None, data_riferimento=None, lat=None, lng=None):
    """Inserisce una nuova notifica per l'admin."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        '''INSERT INTO notifiche (messaggio, tipo, creato_il, user_id, data_riferimento, lat, lng)
           VALUES (?, ?, ?, ?, ?, ?, ?)''',
        (messaggio, tipo, datetime.now().isoformat(), user_id, data_riferimento, lat, lng)
    )
    conn.commit()
    conn.close()


def formatta_elenco_nomi(nomi):
    """Trasforma una lista di nomi in un elenco leggibile: 'A', 'A e B', 'A, B e C'."""
    nomi = list(nomi)
    if not nomi:
        return ''
    if len(nomi) == 1:
        return nomi[0]
    return f"{', '.join(nomi[:-1])} e {nomi[-1]}"


def trova_colleghi_sovrapposti(cursor, user_id, data_inizio, data_fine):
    """Cerca altri dipendenti con ferie/permessi (approvati o in attesa) che si
    sovrappongono, anche solo per un giorno, al periodo indicato. Restituisce
    una lista di nomi completi, senza duplicati, nell'ordine di prima comparsa."""
    cursor.execute(
        '''SELECT richieste.user_id, users.name, users.cognome
           FROM richieste JOIN users ON users.id = richieste.user_id
           WHERE richieste.user_id != ?
             AND richieste.stato IN ('approvata', 'in_attesa')
             AND richieste.data_inizio <= ?
             AND richieste.data_fine >= ?
           ORDER BY richieste.data_inizio ASC''',
        (user_id, data_fine, data_inizio)
    )
    righe = cursor.fetchall()

    colleghi = {}
    for r in righe:
        colleghi.setdefault(r['user_id'], f"{r['name']} {r['cognome']}")
    return list(colleghi.values())


def trova_conflitto_ferie_permesso(cursor, user_id, tipo, data_inizio, data_fine,
                                    escludi_richiesta_id=None, stati=('approvata', 'in_attesa')):
    """True se lo stesso dipendente ha già un'altra richiesta del tipo
    OPPOSTO (ferie se questa è un permesso, permesso se questa è una ferie)
    che si sovrappone, anche solo per un giorno, al periodo indicato: ferie
    e permesso sono mutuamente esclusivi nello stesso giorno per la stessa
    persona."""
    tipo_opposto = 'permesso' if tipo == 'ferie' else 'ferie'
    segnaposto_stati = ','.join('?' for _ in stati)
    query = f'''SELECT 1 FROM richieste
                WHERE user_id = ? AND tipo = ? AND stato IN ({segnaposto_stati})
                  AND data_inizio <= ? AND data_fine >= ?'''
    parametri = [user_id, tipo_opposto, *stati, data_fine, data_inizio]

    if escludi_richiesta_id is not None:
        query += ' AND id != ?'
        parametri.append(escludi_richiesta_id)

    cursor.execute(query, parametri)
    return cursor.fetchone() is not None


def crea_messaggio(user_id, mittente, testo):
    """Inserisce un messaggio nella chat tra admin e dipendente.
    mittente è 'admin' oppure 'utente'. Il messaggio nasce già letto
    da chi lo scrive e da leggere per l'altra parte."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        '''INSERT INTO messaggi (user_id, mittente, testo, creato_il, letto_admin, letto_utente)
           VALUES (?, ?, ?, ?, ?, ?)''',
        (
            user_id,
            mittente,
            testo,
            datetime.now().isoformat(),
            1 if mittente == 'admin' else 0,
            1 if mittente == 'utente' else 0,
        )
    )
    conn.commit()
    conn.close()


def get_orario_giorno(user, data_iso):
    """Restituisce l'orario previsto per uno specifico giorno: se l'admin ha
    impostato una modifica giornaliera per quella data la usa (campo per campo,
    altrimenti ricade sull'orario standard dell'utente), sennò usa l'orario
    standard dell'utente per intero."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        'SELECT * FROM orari_giornalieri WHERE user_id = ? AND data = ?',
        (user['id'], data_iso)
    )
    override = cursor.fetchone()
    conn.close()

    if override:
        return {
            'orario_inizio': override['orario_inizio'] or user['orario_inizio'],
            'orario_pausa_inizio': override['orario_pausa_inizio'] or user['orario_pausa_inizio'],
            'orario_pausa_fine': override['orario_pausa_fine'] or user['orario_pausa_fine'],
            'orario_fine': override['orario_fine'] or user['orario_fine'],
            'permesso_inizio': override['permesso_inizio'],
            'permesso_fine': override['permesso_fine'],
            'is_override': True,
        }

    return {
        'orario_inizio': user['orario_inizio'],
        'orario_pausa_inizio': user['orario_pausa_inizio'],
        'orario_pausa_fine': user['orario_pausa_fine'],
        'orario_fine': user['orario_fine'],
        'permesso_inizio': None,
        'permesso_fine': None,
        'is_override': False,
    }


def ha_permesso_approvato(user_id, data_iso):
    """True se per quella data il dipendente ha una richiesta di ferie o
    permesso già approvata (quindi non va segnalato come assente o come
    monte ore non rispettato)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        '''SELECT 1 FROM richieste
           WHERE user_id = ? AND stato = 'approvata'
             AND data_inizio <= ? AND data_fine >= ?''',
        (user_id, data_iso, data_iso)
    )
    trovato = cursor.fetchone()
    conn.close()
    return trovato is not None


def ha_ferie_approvata(user_id, data_iso):
    """True se per quella data il dipendente ha una richiesta di ferie
    (giornata intera) già approvata: in tal caso non può timbrare."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        '''SELECT 1 FROM richieste
           WHERE user_id = ? AND tipo = 'ferie' AND stato = 'approvata'
             AND data_inizio <= ? AND data_fine >= ?''',
        (user_id, data_iso, data_iso)
    )
    trovato = cursor.fetchone()
    conn.close()
    return trovato is not None


def controlla_assenze():
    """Segnala i dipendenti che non hanno ancora timbrato l'entrata 30 minuti
    dopo il loro orario previsto. Evita duplicati controllando la tabella
    assenze_segnalate, che resta valida anche se l'admin
    cancella la notifica visibile nel pannello."""
    oggi = date.today()
    oggi_iso = oggi.isoformat()
    adesso = datetime.now()

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE orario_inizio IS NOT NULL AND orario_inizio != ''")
    utenti = cursor.fetchall()

    for utente in utenti:
        if ha_permesso_approvato(utente['id'], oggi_iso):
            continue

        orario_giorno = get_orario_giorno(utente, oggi_iso)
        if not orario_giorno['orario_inizio']:
            continue

        try:
            atteso = datetime.strptime(orario_giorno['orario_inizio'], '%H:%M').time()
        except (ValueError, TypeError):
            continue

        atteso_dt = datetime.combine(oggi, atteso)
        if adesso < atteso_dt + timedelta(minutes=30):
            continue

        cursor.execute(
            "SELECT 1 FROM timbrature WHERE user_id = ? AND tipo = 'entrata' AND creato_il LIKE ?",
            (utente['id'], f'{oggi_iso}%')
        )
        if cursor.fetchone():
            continue

        cursor.execute(
            "SELECT 1 FROM assenze_segnalate WHERE user_id = ? AND data_riferimento = ?",
            (utente['id'], oggi_iso)
        )
        if cursor.fetchone():
            continue

        nome_completo = f"{utente['name']} {utente['cognome']}"

        cursor.execute(
            '''INSERT INTO notifiche (messaggio, tipo, creato_il, user_id, data_riferimento)
               VALUES (?, ?, ?, ?, ?)''',
            (
                f"{nome_completo} non ha ancora timbrato l'entrata (prevista alle {utente['orario_inizio']})",
                'assenza',
                adesso.isoformat(),
                utente['id'],
                oggi_iso
            )
        )

        try:
            cursor.execute(
                '''INSERT INTO assenze_segnalate (user_id, data_riferimento, creato_il)
                   VALUES (?, ?, ?)''',
                (utente['id'], oggi_iso, adesso.isoformat())
            )
        except sqlite3.IntegrityError:
            pass

    conn.commit()
    conn.close()


def calcola_stato_giornata(punches, now):
    """Dato un elenco di timbrature di UN SOLO giorno,
    calcola lo stato corrente ('non_iniziato' | 'in_corso' | 'in_pausa' | 'terminato')
    e i secondi complessivamente lavorati, includendo il segmento in corso se presente."""
    stato = 'non_iniziato'
    secondi_lavorati = 0
    inizio_segmento = None

    for p in punches:
        tipo = p['tipo']
        ts = datetime.fromisoformat(p['creato_il'])

        if tipo == 'entrata':
            inizio_segmento = ts
            stato = 'in_corso'
        elif tipo == 'pausa_inizio':
            if inizio_segmento is not None:
                secondi_lavorati += (ts - inizio_segmento).total_seconds()
                inizio_segmento = None
            stato = 'in_pausa'
        elif tipo == 'pausa_fine':
            inizio_segmento = ts
            stato = 'in_corso'
        elif tipo == 'uscita':
            if inizio_segmento is not None:
                secondi_lavorati += (ts - inizio_segmento).total_seconds()
                inizio_segmento = None
            stato = 'terminato'

    if stato == 'in_corso' and inizio_segmento is not None:
        secondi_lavorati += (now - inizio_segmento).total_seconds()

    return stato, max(0, int(secondi_lavorati))


TRANSIZIONI_AMMESSE = {
    'non_iniziato': {'entrata'},
    'in_corso': {'pausa_inizio', 'uscita'},
    'in_pausa': {'pausa_fine'},
    'terminato': set(),
}


def get_dashboard_oggi():
    """Restituisce, per il widget dashboard del pannello admin, l'elenco dei
    dipendenti attivi oggi (cioè non in ferie), di chi è in ferie e di chi è
    in permesso, con lo stato di timbratura per gli attivi."""
    oggi_iso = date.today().isoformat()
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT id, name, cognome FROM users ORDER BY cognome, name')
    tutti_utenti = cursor.fetchall()

    cursor.execute(
        '''SELECT users.id, users.name, users.cognome
           FROM richieste JOIN users ON users.id = richieste.user_id
           WHERE richieste.tipo = 'ferie' AND richieste.stato = 'approvata'
             AND richieste.data_inizio <= ? AND richieste.data_fine >= ?
           ORDER BY users.cognome, users.name''',
        (oggi_iso, oggi_iso)
    )
    in_ferie = cursor.fetchall()
    id_in_ferie = {r['id'] for r in in_ferie}

    cursor.execute(
        '''SELECT users.id, users.name, users.cognome, richieste.ora_inizio, richieste.ora_fine
           FROM richieste JOIN users ON users.id = richieste.user_id
           WHERE richieste.tipo = 'permesso' AND richieste.stato = 'approvata'
             AND richieste.data_inizio <= ? AND richieste.data_fine >= ?
           ORDER BY users.cognome, users.name''',
        (oggi_iso, oggi_iso)
    )
    in_permesso = cursor.fetchall()
    id_in_permesso = {r['id'] for r in in_permesso}

    cursor.execute(
        "SELECT DISTINCT user_id FROM timbrature WHERE tipo = 'entrata' AND creato_il LIKE ?",
        (f'{oggi_iso}%',)
    )
    id_firmati = {r['user_id'] for r in cursor.fetchall()}
    conn.close()

    return {
        'attivi': [
            {
                'id': u['id'],
                'nome': u['name'],
                'cognome': u['cognome'],
                'firmato': u['id'] in id_firmati,
                'in_permesso': u['id'] in id_in_permesso,
            }
            for u in tutti_utenti if u['id'] not in id_in_ferie
        ],
        'ferie': [{'id': r['id'], 'nome': r['name'], 'cognome': r['cognome']} for r in in_ferie],
        'permesso': [
            {
                'id': r['id'], 'nome': r['name'], 'cognome': r['cognome'],
                'ora_inizio': r['ora_inizio'], 'ora_fine': r['ora_fine'],
            }
            for r in in_permesso
        ],
    }


GIORNI_SETTIMANA_IT = ['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì', 'Sabato', 'Domenica']


def leggi_periodo_da_richiesta():
    """Legge i parametri 'da' e 'a' dalla query string per gli export Excel;
    se assenti o non validi usa il mese corrente (dal giorno 1 a oggi)."""
    oggi = date.today()
    da_raw = request.args.get('da', '').strip()
    a_raw = request.args.get('a', '').strip()

    try:
        da = date.fromisoformat(da_raw) if da_raw else oggi.replace(day=1)
    except ValueError:
        da = oggi.replace(day=1)

    try:
        a = date.fromisoformat(a_raw) if a_raw else oggi
    except ValueError:
        a = oggi

    if a < da:
        da, a = a, da

    return da, a


def giorni_intersezione(data_inizio_iso, data_fine_iso, da, a):
    """Numero di giorni di [data_inizio, data_fine] che ricadono anche
    dentro il periodo [da, a] (entrambi gli intervalli inclusivi)."""
    inizio = max(date.fromisoformat(data_inizio_iso), da)
    fine = min(date.fromisoformat(data_fine_iso), a)
    if fine < inizio:
        return 0
    return (fine - inizio).days + 1


def calcola_giornata_per_export(punches, giorno, oggi):
    """Come calcola_stato_giornata, ma pensato per un export su un giorno
    già trascorso: se la giornata risulta incompleta (l'ultima timbratura
    non è un'uscita) non inventa ore lavorate fino ad ora, ma si ferma
    all'ultima timbratura registrata quel giorno."""
    if giorno == oggi:
        riferimento = datetime.now()
    elif punches:
        riferimento = datetime.fromisoformat(punches[-1]['creato_il'])
    else:
        riferimento = datetime.combine(giorno, datetime.min.time())
    return calcola_stato_giornata(punches, riferimento)


def formatta_ore_minuti(secondi):
    ore, resto = divmod(int(secondi), 3600)
    minuti = resto // 60
    return f"{ore}:{minuti:02d}"


def formatta_data_it(d):
    return d.strftime('%d/%m/%Y')


COLORE_INTESTAZIONE_EXCEL = '00407D'


def crea_risposta_excel(titolo, intestazioni, righe, nome_file, righe_extra=None):
    """Costruisce un file .xlsx con una riga di titolo, un'intestazione in
    grassetto e le righe di dati; le colonne vengono allargate in base al
    contenuto più lungo così il testo si legge senza dover ridimensionare
    manualmente le celle. Le date passate come oggetti date vengono scritte
    come date vere (non testo), formattate in stile italiano GG/MM/AAAA.
    righe_extra (opzionale) è un elenco di righe aggiuntive (es. un
    riepilogo statistico) aggiunte sotto una riga vuota, senza una propria
    intestazione."""
    numero_colonne = len(intestazioni)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Dati'

    ws.append([titolo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(numero_colonne, 1))
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)

    ws.append([])

    riga_intestazioni = 3
    ws.append(intestazioni)
    for colonna in range(1, numero_colonne + 1):
        cella = ws.cell(row=riga_intestazioni, column=colonna)
        cella.font = Font(bold=True, color='FFFFFF')
        cella.fill = PatternFill('solid', fgColor=COLORE_INTESTAZIONE_EXCEL)
        cella.alignment = Alignment(vertical='center')

    prima_riga_dati = riga_intestazioni + 1
    for riga in righe:
        ws.append(riga)

    ultima_riga_dati = prima_riga_dati + len(righe) - 1
    for indice_riga in range(prima_riga_dati, ultima_riga_dati + 1):
        for colonna in range(1, numero_colonne + 1):
            cella = ws.cell(row=indice_riga, column=colonna)
            if isinstance(cella.value, date):
                cella.number_format = 'DD/MM/YYYY'

    if righe_extra:
        ws.append([])
        riga_titolo_extra = ws.max_row + 1
        for riga in righe_extra:
            ws.append(riga)
        ws.cell(row=riga_titolo_extra, column=1).font = Font(bold=True)

    larghezze = [len(str(intestazione)) for intestazione in intestazioni]
    for riga in righe:
        for indice, valore in enumerate(riga):
            testo = formatta_data_it(valore) if isinstance(valore, date) else str(valore)
            larghezze[indice] = max(larghezze[indice], len(testo))
    if righe_extra:
        for riga in righe_extra:
            for indice, valore in enumerate(riga):
                if indice < len(larghezze):
                    larghezze[indice] = max(larghezze[indice], len(str(valore)))

    for indice, larghezza in enumerate(larghezze, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = min(max(larghezza + 4, 12), 45)

    ws.freeze_panes = f'A{prima_riga_dati}'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nome_file}"'}
    )


def require_admin():
    """Restituisce None se l'utente è admin, altrimenti un redirect al login admin."""
    if session.get('tipo_utente') != 'admin':
        return redirect(url_for('admin_login_page'))
    return None


@app.route('/')
def index():
    if 'user' in session:
        if session.get('tipo_utente') == 'admin':
            return redirect(url_for('admin_panel'))
        return redirect(url_for('timbro_page'))
    return render_template('index.html')


@app.route('/privacy', methods=['GET'])
def privacy_page():
    if 'user' in session:
        if session.get('tipo_utente') == 'admin':
            link_indietro, etichetta_indietro = url_for('admin_panel'), '← Torna al pannello'
        else:
            link_indietro, etichetta_indietro = url_for('timbro_page'), '← Torna alla timbratura'
    else:
        link_indietro, etichetta_indietro = url_for('index'), '← Torna alla home'
    return render_template('privacy.html', link_indietro=link_indietro, etichetta_indietro=etichetta_indietro)


@app.route('/admin/login', methods=['GET'])
def admin_login_page():
    return render_template('admin_login.html')


@app.route('/login', methods=['GET'])
def login_page():
    return render_template('login.html')


@app.route('/admin/login', methods=['POST'])
def admin_login():
    nome = request.form.get('nome', '').strip()
    password = request.form.get('password', '').strip()

    if nome == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        session.permanent = True
        session['user'] = {
            'email': 'admin@gmail.com',
            'name': 'Administrator',
            'authenticated_at': datetime.utcnow().isoformat()
        }
        session['tipo_utente'] = 'admin'
        return redirect(url_for('admin_panel'))

    return jsonify({'ok': False, 'errore': 'Credenziali non valide'}), 401


@app.route('/admin/panel')
def admin_panel():
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    controlla_assenze()

    dashboard_oggi = get_dashboard_oggi()
    dashboard_conteggi = {
        'attivi': len(dashboard_oggi['attivi']),
        'firmati': sum(1 for u in dashboard_oggi['attivi'] if u['firmato']),
        'ferie': len(dashboard_oggi['ferie']),
        'permesso': len(dashboard_oggi['permesso']),
    }

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT users.*,
            (SELECT COUNT(*) FROM messaggi
             WHERE messaggi.user_id = users.id
               AND messaggi.mittente = 'utente'
               AND messaggi.letto_admin = 0) AS messaggi_non_letti
        FROM users ORDER BY cognome, name
    ''')
    users = cursor.fetchall()

    cursor.execute('SELECT * FROM notifiche ORDER BY creato_il DESC LIMIT 30')
    notifiche = cursor.fetchall()
    cursor.execute('SELECT COUNT(*) AS n FROM notifiche WHERE letta = 0')
    notifiche_non_lette = cursor.fetchone()['n']

    cursor.execute('''
        SELECT richieste.*, users.name, users.cognome
        FROM richieste JOIN users ON users.id = richieste.user_id
        WHERE richieste.stato = 'in_attesa'
        ORDER BY richieste.creato_il ASC
    ''')
    richieste_in_attesa = cursor.fetchall()

    cursor.execute('''
        SELECT richieste.*, users.name, users.cognome
        FROM richieste JOIN users ON users.id = richieste.user_id
        WHERE richieste.stato != 'in_attesa'
        ORDER BY richieste.gestita_il DESC LIMIT 15
    ''')
    richieste_gestite = cursor.fetchall()

    cursor.execute('''
        SELECT richieste.*, users.name, users.cognome
        FROM richieste JOIN users ON users.id = richieste.user_id
        WHERE richieste.stato = 'approvata' AND richieste.rimozione_richiesta = 1
        ORDER BY richieste.creato_il ASC
    ''')
    richieste_rimozione = cursor.fetchall()
    conn.close()

    return render_template(
        'admin_panel.html',
        admin_name=session['user']['name'],
        users=users,
        notifiche=notifiche,
        notifiche_non_lette=notifiche_non_lette,
        richieste_in_attesa=richieste_in_attesa,
        richieste_gestite=richieste_gestite,
        richieste_rimozione=richieste_rimozione,
        dashboard_oggi=dashboard_oggi,
        dashboard_conteggi=dashboard_conteggi,
        export_da=date.today().replace(day=1).isoformat(),
        export_a=date.today().isoformat()
    )


ORE_MENSILI_MASSIME = 160


def _statistiche_utente_periodo(cursor, utente, da, a):
    """Calcola le statistiche di un dipendente nel periodo [da, a]: giorni
    lavorati, ore lavorate, giorni di ferie, ore di permesso, conteggio di
    ritardi/assenze segnalati, e ore di straordinario/ore mancanti rispetto
    al tetto massimo di ORE_MENSILI_MASSIME ore lavorate nel periodo."""
    user_id = utente['id']
    da_iso, a_iso = da.isoformat(), a.isoformat()
    oggi = date.today()

    cursor.execute(
        "SELECT tipo, creato_il FROM timbrature WHERE user_id = ? AND creato_il >= ? AND creato_il < ? ORDER BY creato_il ASC",
        (user_id, da_iso, (a + timedelta(days=1)).isoformat())
    )
    per_giorno = {}
    for r in cursor.fetchall():
        per_giorno.setdefault(r['creato_il'][:10], []).append(dict(r))

    giorni_lavorati = 0
    secondi_totali = 0
    for giorno_str, punches in per_giorno.items():
        _, secondi = calcola_giornata_per_export(punches, date.fromisoformat(giorno_str), oggi)
        if secondi > 0:
            giorni_lavorati += 1
        secondi_totali += secondi

    cursor.execute(
        '''SELECT data_inizio, data_fine FROM richieste
           WHERE user_id = ? AND tipo = 'ferie' AND stato = 'approvata'
             AND data_inizio <= ? AND data_fine >= ?''',
        (user_id, a_iso, da_iso)
    )
    giorni_ferie = sum(giorni_intersezione(r['data_inizio'], r['data_fine'], da, a) for r in cursor.fetchall())

    cursor.execute(
        '''SELECT ora_inizio, ora_fine FROM richieste
           WHERE user_id = ? AND tipo = 'permesso' AND stato = 'approvata'
             AND data_inizio <= ? AND data_fine >= ?''',
        (user_id, a_iso, da_iso)
    )
    secondi_permesso = 0
    for r in cursor.fetchall():
        if r['ora_inizio'] and r['ora_fine']:
            try:
                inizio_dt = datetime.strptime(r['ora_inizio'], '%H:%M')
                fine_dt = datetime.strptime(r['ora_fine'], '%H:%M')
                secondi_permesso += (fine_dt - inizio_dt).total_seconds()
            except ValueError:
                pass

    cursor.execute(
        '''SELECT tipo, COUNT(*) AS n FROM notifiche
           WHERE user_id = ? AND tipo IN ('alert', 'assenza')
             AND creato_il >= ? AND creato_il < ?
           GROUP BY tipo''',
        (user_id, da_iso, (a + timedelta(days=1)).isoformat())
    )
    conteggi = {r['tipo']: r['n'] for r in cursor.fetchall()}

    secondi_target = ORE_MENSILI_MASSIME * 3600
    secondi_straordinario = max(0, secondi_totali - secondi_target)
    secondi_mancanti = max(0, secondi_target - secondi_totali)

    return {
        'giorni_lavorati': giorni_lavorati,
        'secondi_lavorati': secondi_totali,
        'giorni_ferie': giorni_ferie,
        'secondi_permesso': secondi_permesso,
        'ritardi': conteggi.get('alert', 0),
        'assenze': conteggi.get('assenza', 0),
        'secondi_straordinario': secondi_straordinario,
        'secondi_mancanti': secondi_mancanti,
    }


@app.route('/admin/export/riepilogo')
def export_riepilogo_excel():
    """Esporta in Excel, per tutti i dipendenti, le statistiche del periodo
    indicato (query string 'da'/'a', default: mese corrente)."""
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    da, a = leggi_periodo_da_richiesta()

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users ORDER BY cognome, name')
    utenti = cursor.fetchall()

    intestazioni = ['Nome', 'Cognome', 'Email', 'Giorni lavorati', 'Ore lavorate',
                    'Giorni di ferie', 'Ore di permesso', 'Ritardi',
                    f'Ore di straordinario (oltre {ORE_MENSILI_MASSIME}h)',
                    f'Ore mancanti (su {ORE_MENSILI_MASSIME}h)', 'Assenze segnalate']

    righe = []
    for utente in utenti:
        s = _statistiche_utente_periodo(cursor, utente, da, a)
        righe.append([
            utente['name'], utente['cognome'], utente['email'],
            s['giorni_lavorati'],
            formatta_ore_minuti(s['secondi_lavorati']),
            s['giorni_ferie'],
            formatta_ore_minuti(s['secondi_permesso']),
            s['ritardi'],
            formatta_ore_minuti(s['secondi_straordinario']),
            formatta_ore_minuti(s['secondi_mancanti']),
            s['assenze'],
        ])

    conn.close()

    titolo = f"Riepilogo dipendenti dal {formatta_data_it(da)} al {formatta_data_it(a)}"
    nome_file = f"riepilogo_{da.isoformat()}_{a.isoformat()}.xlsx"
    return crea_risposta_excel(titolo, intestazioni, righe, nome_file)


@app.route('/admin/export/utente/<int:user_id>')
def export_utente_excel(user_id):
    """Esporta in Excel il dettaglio giornaliero (senza sabati e domeniche)
    di un dipendente nel periodo indicato (query string 'da'/'a', default:
    mese corrente), con un riepilogo statistico in coda."""
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    utente = cursor.fetchone()
    if utente is None:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Utente non trovato'}), 404

    da, a = leggi_periodo_da_richiesta()
    da_iso, a_iso = da.isoformat(), a.isoformat()
    oggi = date.today()

    cursor.execute(
        "SELECT tipo, creato_il FROM timbrature WHERE user_id = ? AND creato_il >= ? AND creato_il < ? ORDER BY creato_il ASC",
        (user_id, da_iso, (a + timedelta(days=1)).isoformat())
    )
    per_giorno = {}
    for r in cursor.fetchall():
        per_giorno.setdefault(r['creato_il'][:10], []).append(dict(r))

    cursor.execute(
        'SELECT * FROM orari_giornalieri WHERE user_id = ? AND data >= ? AND data <= ?',
        (user_id, da_iso, a_iso)
    )
    override_per_giorno = {r['data']: r for r in cursor.fetchall()}

    statistiche = _statistiche_utente_periodo(cursor, utente, da, a)
    conn.close()

    intestazioni = ['Data', 'Giorno', 'Entrata', 'Inizio pausa', 'Fine pausa', 'Uscita',
                    'Ore lavorate', 'Ferie', 'Permesso', 'Nota']

    righe = []
    giorno = da
    while giorno <= a:
        if giorno.weekday() >= 5:  # sabato (5) e domenica (6): non giornate lavorative
            giorno += timedelta(days=1)
            continue

        giorno_iso = giorno.isoformat()
        punches = per_giorno.get(giorno_iso, [])
        etichette_punch = {p['tipo']: datetime.fromisoformat(p['creato_il']).strftime('%H:%M') for p in punches}
        _, secondi = calcola_giornata_per_export(punches, giorno, oggi)

        override = override_per_giorno.get(giorno_iso)
        permesso_str = ''
        if override and override['permesso_inizio'] and override['permesso_fine']:
            permesso_str = f"{override['permesso_inizio']}–{override['permesso_fine']}"

        righe.append([
            giorno,
            GIORNI_SETTIMANA_IT[giorno.weekday()],
            etichette_punch.get('entrata', ''),
            etichette_punch.get('pausa_inizio', ''),
            etichette_punch.get('pausa_fine', ''),
            etichette_punch.get('uscita', ''),
            formatta_ore_minuti(secondi) if secondi > 0 else '',
            'Sì' if ha_ferie_approvata(user_id, giorno_iso) else '',
            permesso_str,
            (override['note'] if override else '') or '',
        ])
        giorno += timedelta(days=1)

    righe_extra = [
        ['Riepilogo periodo'],
        ['Giorni lavorati', statistiche['giorni_lavorati']],
        ['Ore lavorate totali', formatta_ore_minuti(statistiche['secondi_lavorati'])],
        ['Giorni di ferie', statistiche['giorni_ferie']],
        ['Ore di permesso', formatta_ore_minuti(statistiche['secondi_permesso'])],
        ['Ritardi', statistiche['ritardi']],
        [f'Ore di straordinario (oltre {ORE_MENSILI_MASSIME}h)', formatta_ore_minuti(statistiche['secondi_straordinario'])],
        [f'Ore mancanti (su {ORE_MENSILI_MASSIME}h)', formatta_ore_minuti(statistiche['secondi_mancanti'])],
        ['Assenze segnalate', statistiche['assenze']],
    ]

    titolo = f"Riepilogo di {utente['name']} {utente['cognome']} ({utente['email']}) dal {formatta_data_it(da)} al {formatta_data_it(a)}"
    nome_file = f"{utente['cognome']}_{utente['name']}_{da_iso}_{a_iso}.xlsx".replace(' ', '_')
    return crea_risposta_excel(titolo, intestazioni, righe, nome_file, righe_extra=righe_extra)


@app.route('/admin/create_user', methods=['GET'])
def create_user_page():
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response
    return render_template('admin_crea.html')


@app.route('/admin/create_user', methods=['POST'])
def create_user():
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    nome = request.form.get('nome', '').strip()
    cognome = request.form.get('cognome', '').strip()
    orario_inizio = request.form.get('orario_inizio', '').strip()
    orario_pausa_inizio = request.form.get('orario_pausa_inizio', '').strip()
    orario_pausa_fine = request.form.get('orario_pausa_fine', '').strip()
    orario_fine = request.form.get('orario_fine', '').strip()
    email = request.form.get('email', '').strip().lower()

    if not nome or not cognome or not email:
        return jsonify({
            'ok': False,
            'errore': 'Nome, cognome ed email sono obbligatori'
        }), 400

    password_provvisoria = secrets.token_urlsafe(8)
    password_hash = generate_password_hash(password_provvisoria)

    conn = get_db()
    cursor = conn.cursor()

    try:
        cursor.execute(
            '''
            INSERT INTO users
            (email, password, name, cognome,
             orario_inizio, orario_pausa_inizio,
             orario_pausa_fine, orario_fine)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                email,
                password_hash,
                nome,
                cognome,
                orario_inizio,
                orario_pausa_inizio,
                orario_pausa_fine,
                orario_fine
            )
        )

        conn.commit()

    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({
            'ok': False,
            'errore': 'Email già registrata'
        }), 409

    conn.close()

    corpo_email = f"""
Ciao {nome},

il tuo account è stato creato.

Email: {email}

Password provvisoria:
{password_provvisoria}

Ti consigliamo di cambiarla al primo accesso.
"""

    try:
        invia_email(email, 'Password provvisoria per il tuo account', corpo_email)
    except Exception as e:
        return jsonify({
            'ok': False,
            'errore': f'Utente creato ma impossibile inviare l\'email: {str(e)}'
        }), 500

    return render_template(
        'admin_crea_ok.html',
        email=email,
        password=password_provvisoria
    )


@app.route('/admin/modifica/<int:user_id>', methods=['GET'])
def edit_user_page(user_id):
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()

    if user is None:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Utente non trovato'}), 404

    cursor.execute(
        '''SELECT * FROM orari_giornalieri
           WHERE user_id = ? AND data >= ?
           ORDER BY data ASC''',
        (user_id, date.today().isoformat())
    )
    orari_giornalieri = cursor.fetchall()
    conn.close()

    return render_template(
        'admin_modifica.html',
        user=user,
        orari_giornalieri=orari_giornalieri,
        oggi=date.today().isoformat(),
        export_da=date.today().replace(day=1).isoformat(),
        export_a=date.today().isoformat()
    )


@app.route('/admin/modifica/<int:user_id>', methods=['POST'])
def edit_user(user_id):
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    nome = request.form.get('nome', '').strip()
    cognome = request.form.get('cognome', '').strip()
    email = request.form.get('email', '').strip().lower()
    orario_inizio = request.form.get('orario_inizio', '').strip()
    orario_pausa_inizio = request.form.get('orario_pausa_inizio', '').strip()
    orario_pausa_fine = request.form.get('orario_pausa_fine', '').strip()
    orario_fine = request.form.get('orario_fine', '').strip()

    if not nome or not cognome or not email:
        return jsonify({'ok': False, 'errore': 'Nome, cognome ed email sono obbligatori'}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            '''UPDATE users
               SET name = ?, cognome = ?, email = ?, orario_inizio = ?,
                   orario_pausa_inizio = ?, orario_pausa_fine = ?, orario_fine = ?
               WHERE id = ?''',
            (nome, cognome, email, orario_inizio, orario_pausa_inizio,
             orario_pausa_fine, orario_fine, user_id)
        )
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'ok': False, 'errore': 'Utente non trovato'}), 404
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Email già registrata da un altro utente'}), 409
    conn.close()

    return redirect(url_for('admin_panel'))


@app.route('/admin/orario_giornaliero/<int:user_id>', methods=['POST'])
def crea_orario_giornaliero(user_id):
    """Permette all'admin di impostare un orario diverso da quello standard
    per un singolo giorno di un dipendente (es. turno straordinario,
    orario ridotto, uscita anticipata concordata, ecc.)."""
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT 1 FROM users WHERE id = ?', (user_id,))
    if cursor.fetchone() is None:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Utente non trovato'}), 404

    data_riferimento = request.form.get('data', '').strip()
    orario_inizio = request.form.get('orario_inizio', '').strip() or None
    orario_pausa_inizio = request.form.get('orario_pausa_inizio', '').strip() or None
    orario_pausa_fine = request.form.get('orario_pausa_fine', '').strip() or None
    orario_fine = request.form.get('orario_fine', '').strip() or None
    note = request.form.get('note', '').strip() or None

    if not data_riferimento:
        conn.close()
        return jsonify({'ok': False, 'errore': 'La data è obbligatoria'}), 400

    cursor.execute(
        '''INSERT INTO orari_giornalieri
           (user_id, data, orario_inizio, orario_pausa_inizio, orario_pausa_fine, orario_fine, note, creato_il)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(user_id, data) DO UPDATE SET
             orario_inizio = excluded.orario_inizio,
             orario_pausa_inizio = excluded.orario_pausa_inizio,
             orario_pausa_fine = excluded.orario_pausa_fine,
             orario_fine = excluded.orario_fine,
             note = excluded.note''',
        (
            user_id, data_riferimento, orario_inizio, orario_pausa_inizio,
            orario_pausa_fine, orario_fine, note, datetime.now().isoformat()
        )
    )
    conn.commit()
    conn.close()

    return redirect(url_for('edit_user_page', user_id=user_id))


@app.route('/admin/orario_giornaliero/<int:user_id>/elimina/<data_riferimento>', methods=['POST'])
def elimina_orario_giornaliero(user_id, data_riferimento):
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        'DELETE FROM orari_giornalieri WHERE user_id = ? AND data = ?',
        (user_id, data_riferimento)
    )
    conn.commit()
    conn.close()

    return redirect(url_for('edit_user_page', user_id=user_id))


@app.route('/admin/elimina/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    eliminato = cursor.rowcount > 0
    conn.close()

    if not eliminato:
        return jsonify({'ok': False, 'errore': 'Utente non trovato'}), 404

    return redirect(url_for('admin_panel'))


@app.route('/admin/richieste/<int:richiesta_id>/approva', methods=['POST'])
def approva_richiesta(richiesta_id):
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response
    return _gestisci_richiesta(richiesta_id, 'approvata')


@app.route('/admin/richieste/<int:richiesta_id>/rifiuta', methods=['POST'])
def rifiuta_richiesta(richiesta_id):
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response
    return _gestisci_richiesta(richiesta_id, 'rifiutata')


def _gestisci_richiesta(richiesta_id, nuovo_stato):
    """Approva o rifiuta una richiesta di ferie/permesso e avvisa il
    dipendente in chat dell'esito."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM richieste WHERE id = ?', (richiesta_id,))
    richiesta = cursor.fetchone()

    if richiesta is None:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Richiesta non trovata'}), 404

    if nuovo_stato == 'approvata' and trova_conflitto_ferie_permesso(
            cursor, richiesta['user_id'], richiesta['tipo'],
            richiesta['data_inizio'], richiesta['data_fine'],
            escludi_richiesta_id=richiesta_id, stati=('approvata',)):
        conn.close()
        tipo_opposto = 'permesso' if richiesta['tipo'] == 'ferie' else 'ferie'
        return jsonify({
            'ok': False,
            'errore': f"Il dipendente ha già una richiesta di {tipo_opposto} approvata che si sovrappone a questo "
                      f"periodo: ferie e permesso non possono coesistere nello stesso giorno"
        }), 409

    cursor.execute(
        'UPDATE richieste SET stato = ?, gestita_il = ? WHERE id = ?',
        (nuovo_stato, datetime.now().isoformat(), richiesta_id)
    )

    if (nuovo_stato == 'approvata' and richiesta['tipo'] == 'permesso'
            and richiesta['ora_inizio'] and richiesta['ora_fine']):
        cursor.execute('SELECT * FROM users WHERE id = ?', (richiesta['user_id'],))
        dipendente = cursor.fetchone()
        orario_std_inizio = dipendente['orario_inizio'] if dipendente else None
        orario_std_fine = dipendente['orario_fine'] if dipendente else None

        giorno = date.fromisoformat(richiesta['data_inizio'])
        fine = date.fromisoformat(richiesta['data_fine'])
        nota = f"Permesso (inizio alle {richiesta['ora_inizio']} e fine alle {richiesta['ora_fine']})"
        while giorno <= fine:
            data_iso = giorno.isoformat()
            cursor.execute(
                '''INSERT INTO orari_giornalieri
                   (user_id, data, orario_inizio, orario_fine, permesso_inizio, permesso_fine, note, creato_il)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(user_id, data) DO UPDATE SET
                     orario_inizio = COALESCE(orari_giornalieri.orario_inizio, excluded.orario_inizio),
                     orario_fine = COALESCE(orari_giornalieri.orario_fine, excluded.orario_fine),
                     permesso_inizio = excluded.permesso_inizio,
                     permesso_fine = excluded.permesso_fine,
                     note = excluded.note''',
                (richiesta['user_id'], data_iso, orario_std_inizio, orario_std_fine,
                 richiesta['ora_inizio'], richiesta['ora_fine'], nota, datetime.now().isoformat())
            )
            giorno += timedelta(days=1)

    conn.commit()
    conn.close()

    etichetta_tipo = 'ferie' if richiesta['tipo'] == 'ferie' else 'permesso'
    if richiesta['data_inizio'] == richiesta['data_fine']:
        periodo = richiesta['data_inizio']
    else:
        periodo = f"dal {richiesta['data_inizio']} al {richiesta['data_fine']}"

    if nuovo_stato == 'approvata':
        testo_messaggio = f"La tua richiesta di {etichetta_tipo} ({periodo}) è stata approvata."
    else:
        testo_messaggio = f"La tua richiesta di {etichetta_tipo} ({periodo}) è stata rifiutata."
    crea_messaggio(richiesta['user_id'], 'admin', testo_messaggio)

    return redirect(url_for('admin_panel'))


@app.route('/admin/richieste/<int:richiesta_id>/approva_rimozione', methods=['POST'])
def approva_rimozione_richiesta(richiesta_id):
    """L'admin conferma la rimozione di una ferie/permesso già approvato,
    su richiesta del dipendente. La richiesta passa allo stato 'revocata' e,
    per i permessi, viene ripulito l'orario giornaliero che era stato creato
    al momento dell'approvazione."""
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM richieste WHERE id = ?', (richiesta_id,))
    richiesta = cursor.fetchone()

    if richiesta is None:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Richiesta non trovata'}), 404

    if richiesta['stato'] != 'approvata' or not richiesta['rimozione_richiesta']:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Nessuna rimozione da approvare per questa richiesta'}), 400

    cursor.execute(
        "UPDATE richieste SET stato = 'revocata', rimozione_richiesta = 0, gestita_il = ? WHERE id = ?",
        (datetime.now().isoformat(), richiesta_id)
    )

    if richiesta['tipo'] == 'permesso' and richiesta['ora_inizio'] and richiesta['ora_fine']:
        giorno = date.fromisoformat(richiesta['data_inizio'])
        fine = date.fromisoformat(richiesta['data_fine'])
        while giorno <= fine:
            cursor.execute(
                '''UPDATE orari_giornalieri SET permesso_inizio = NULL, permesso_fine = NULL
                   WHERE user_id = ? AND data = ?''',
                (richiesta['user_id'], giorno.isoformat())
            )
            giorno += timedelta(days=1)

    conn.commit()
    conn.close()

    etichetta_tipo = 'ferie' if richiesta['tipo'] == 'ferie' else 'permesso'
    periodo = richiesta['data_inizio'] if richiesta['data_inizio'] == richiesta['data_fine'] else f"dal {richiesta['data_inizio']} al {richiesta['data_fine']}"
    crea_messaggio(
        richiesta['user_id'], 'admin',
        f"La rimozione di {etichetta_tipo} ({periodo}) è stata confermata."
    )

    return redirect(url_for('admin_panel'))


@app.route('/admin/richieste/<int:richiesta_id>/rifiuta_rimozione', methods=['POST'])
def rifiuta_rimozione_richiesta(richiesta_id):
    """L'admin rifiuta la richiesta di rimozione: la ferie/permesso resta
    approvato/a esattamente come prima."""
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM richieste WHERE id = ?', (richiesta_id,))
    richiesta = cursor.fetchone()

    if richiesta is None:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Richiesta non trovata'}), 404

    if richiesta['stato'] != 'approvata' or not richiesta['rimozione_richiesta']:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Nessuna rimozione da rifiutare per questa richiesta'}), 400

    cursor.execute(
        "UPDATE richieste SET rimozione_richiesta = 0 WHERE id = ?",
        (richiesta_id,)
    )
    conn.commit()
    conn.close()

    etichetta_tipo = 'ferie' if richiesta['tipo'] == 'ferie' else 'permesso'
    periodo = richiesta['data_inizio'] if richiesta['data_inizio'] == richiesta['data_fine'] else f"dal {richiesta['data_inizio']} al {richiesta['data_fine']}"
    crea_messaggio(
        richiesta['user_id'], 'admin',
        f"La richiesta di rimozione per {etichetta_tipo} ({periodo}) è stata rifiutata: resta valido/a come approvato/a."
    )

    return redirect(url_for('admin_panel'))


@app.route('/richieste', methods=['GET'])
def richieste_page():
    if 'user' not in session:
        return redirect(url_for('login_page'))
    if session.get('tipo_utente') == 'admin':
        return redirect(url_for('admin_panel'))

    user_id = session['user'].get('id')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    utente = cursor.fetchone()
    cursor.execute(
        'SELECT * FROM richieste WHERE user_id = ? ORDER BY data_inizio DESC',
        (user_id,)
    )
    mie_richieste = cursor.fetchall()
    conn.close()

    return render_template('richieste.html', richieste=mie_richieste, utente=utente)


@app.route('/richieste/verifica_sovrapposizione', methods=['POST'])
def verifica_sovrapposizione_richiesta():
    """Controlla, senza creare nulla, se per il periodo indicato altri
    colleghi hanno già ferie/permessi (approvati o in attesa) che si
    sovrappongono. Usata dal front-end per mostrare un popup di conferma
    PRIMA di inviare effettivamente la richiesta."""
    if 'user' not in session or session.get('tipo_utente') == 'admin':
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    data_inizio = request.form.get('data_inizio', '').strip()
    data_fine = request.form.get('data_fine', '').strip() or data_inizio
    if not data_inizio:
        return jsonify({'ok': False, 'errore': 'La data di inizio è obbligatoria'}), 400

    user_id = session['user'].get('id')
    conn = get_db()
    cursor = conn.cursor()
    colleghi_sovrapposti = trova_colleghi_sovrapposti(cursor, user_id, data_inizio, data_fine)
    conn.close()

    return jsonify({'ok': True, 'colleghi': colleghi_sovrapposti})


@app.route('/richieste/nuova', methods=['POST'])
def crea_richiesta():
    if 'user' not in session or session.get('tipo_utente') == 'admin':
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    tipo = request.form.get('tipo', '').strip()
    data_inizio = request.form.get('data_inizio', '').strip()
    data_fine = request.form.get('data_fine', '').strip() or data_inizio
    motivo = request.form.get('motivo', '').strip()
    ora_inizio = request.form.get('ora_inizio', '').strip() or None
    ora_fine = request.form.get('ora_fine', '').strip() or None

    if tipo not in ('ferie', 'permesso'):
        return jsonify({'ok': False, 'errore': 'Tipo di richiesta non valido'}), 400
    if not data_inizio:
        return jsonify({'ok': False, 'errore': 'La data di inizio è obbligatoria'}), 400
    if data_fine < data_inizio:
        return jsonify({'ok': False, 'errore': 'La data di fine non può precedere quella di inizio'}), 400

    if tipo == 'permesso':
        # Il permesso è orario: ha senso solo per una singola giornata.
        data_fine = data_inizio

        if not ora_inizio or not ora_fine:
            return jsonify({'ok': False, 'errore': "Per il permesso indica l'ora di inizio e di fine"}), 400
        try:
            ora_inizio_dt = datetime.strptime(ora_inizio, '%H:%M')
            ora_fine_dt = datetime.strptime(ora_fine, '%H:%M')
        except ValueError:
            return jsonify({'ok': False, 'errore': 'Orario non valido'}), 400
        if ora_fine_dt <= ora_inizio_dt:
            return jsonify({'ok': False, 'errore': "L'ora di fine deve essere successiva all'ora di inizio"}), 400

        user_id_check = session['user'].get('id')
        conn_check = get_db()
        cursor_check = conn_check.cursor()
        cursor_check.execute('SELECT * FROM users WHERE id = ?', (user_id_check,))
        utente_check = cursor_check.fetchone()
        conn_check.close()

        orario_lavoro = get_orario_giorno(utente_check, data_inizio) if utente_check else None
        if not orario_lavoro or not orario_lavoro['orario_inizio'] or not orario_lavoro['orario_fine']:
            return jsonify({'ok': False, 'errore': "Non è impostato un orario di lavoro per questo giorno: contatta l'amministratore"}), 400

        try:
            lavoro_inizio_dt = datetime.strptime(orario_lavoro['orario_inizio'], '%H:%M')
            lavoro_fine_dt = datetime.strptime(orario_lavoro['orario_fine'], '%H:%M')
        except ValueError:
            return jsonify({'ok': False, 'errore': "Orario di lavoro previsto non valido: contatta l'amministratore"}), 400

        if ora_inizio_dt < lavoro_inizio_dt or ora_fine_dt > lavoro_fine_dt:
            return jsonify({
                'ok': False,
                'errore': f"Il permesso deve rientrare nell'orario di lavoro previsto per quel giorno "
                          f"({orario_lavoro['orario_inizio']}–{orario_lavoro['orario_fine']})"
            }), 400
    else:
        ora_inizio = None
        ora_fine = None

    user_id = session['user'].get('id')

    conn = get_db()
    cursor = conn.cursor()

    if trova_conflitto_ferie_permesso(cursor, user_id, tipo, data_inizio, data_fine):
        conn.close()
        tipo_opposto = 'permesso' if tipo == 'ferie' else 'ferie'
        return jsonify({
            'ok': False,
            'errore': f"Hai già una richiesta di {tipo_opposto} (approvata o in attesa) che si sovrappone a questo "
                      f"periodo: ferie e permesso non possono coesistere nello stesso giorno"
        }), 409

    cursor.execute(
        '''INSERT INTO richieste (user_id, tipo, data_inizio, data_fine, motivo, stato, creato_il, ora_inizio, ora_fine)
           VALUES (?, ?, ?, ?, ?, 'in_attesa', ?, ?, ?)''',
        (user_id, tipo, data_inizio, data_fine, motivo or None, datetime.now().isoformat(), ora_inizio, ora_fine)
    )
    conn.commit()

    cursor.execute('SELECT name, cognome FROM users WHERE id = ?', (user_id,))
    utente = cursor.fetchone()

    colleghi_sovrapposti = trova_colleghi_sovrapposti(cursor, user_id, data_inizio, data_fine)
    conn.close()

    nome_completo = f"{utente['name']} {utente['cognome']}" if utente else 'Un dipendente'
    etichetta_tipo = 'ferie' if tipo == 'ferie' else 'permesso'
    periodo = data_inizio if data_inizio == data_fine else f"dal {data_inizio} al {data_fine}"
    if tipo == 'permesso' and ora_inizio and ora_fine:
        periodo += f", inizio alle {ora_inizio} e fine alle {ora_fine}"

    crea_notifica(
        f"{nome_completo} ha richiesto {etichetta_tipo} ({periodo})",
        tipo='richiesta',
        user_id=user_id,
        data_riferimento=data_inizio
    )

    # Avvisa l'admin se in quel periodo ci sono già altri colleghi con
    # ferie/permessi richiesti o approvati (il dipendente è già stato
    # avvisato ed ha confermato prima dell'invio, tramite popup).
    if colleghi_sovrapposti:
        elenco = formatta_elenco_nomi(colleghi_sovrapposti)
        verbo = 'ha' if len(colleghi_sovrapposti) == 1 else 'hanno'

        crea_notifica(
            f"Più assenze nello stesso periodo: oltre a {nome_completo}, anche {elenco} "
            f"{verbo} ferie o permessi che si sovrappongono ({periodo})",
            tipo='sovrapposizione',
            user_id=user_id,
            data_riferimento=data_inizio
        )

    return redirect(url_for('richieste_page'))


@app.route('/richieste/<int:richiesta_id>/elimina', methods=['POST'])
def elimina_richiesta(richiesta_id):
    """Permette al dipendente di eliminare una propria richiesta ancora in
    attesa di approvazione. Una richiesta già approvata o rifiutata non può
    essere eliminata direttamente: per quelle approvate esiste la richiesta
    di rimozione, che deve passare dall'admin."""
    if 'user' not in session or session.get('tipo_utente') == 'admin':
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    user_id = session['user'].get('id')

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM richieste WHERE id = ? AND user_id = ?', (richiesta_id, user_id))
    richiesta = cursor.fetchone()

    if richiesta is None:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Richiesta non trovata'}), 404

    if richiesta['stato'] != 'in_attesa':
        conn.close()
        return jsonify({'ok': False, 'errore': 'Puoi eliminare solo le richieste ancora in attesa'}), 400

    cursor.execute('DELETE FROM richieste WHERE id = ?', (richiesta_id,))
    conn.commit()
    conn.close()

    return redirect(url_for('richieste_page'))


@app.route('/richieste/<int:richiesta_id>/richiedi_rimozione', methods=['POST'])
def richiedi_rimozione_richiesta(richiesta_id):
    """Permette al dipendente di chiedere all'admin la rimozione di una
    ferie o permesso già approvato. Non elimina nulla direttamente: segnala
    solo la richiesta all'amministratore, che dovrà approvarla."""
    if 'user' not in session or session.get('tipo_utente') == 'admin':
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    user_id = session['user'].get('id')

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM richieste WHERE id = ? AND user_id = ?', (richiesta_id, user_id))
    richiesta = cursor.fetchone()

    if richiesta is None:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Richiesta non trovata'}), 404

    if richiesta['stato'] != 'approvata':
        conn.close()
        return jsonify({'ok': False, 'errore': 'Puoi chiedere la rimozione solo di una richiesta già approvata'}), 400

    if richiesta['rimozione_richiesta']:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Hai già chiesto la rimozione di questa richiesta'}), 400

    cursor.execute('UPDATE richieste SET rimozione_richiesta = 1 WHERE id = ?', (richiesta_id,))

    cursor.execute('SELECT name, cognome FROM users WHERE id = ?', (user_id,))
    utente = cursor.fetchone()
    conn.commit()
    conn.close()

    nome_completo = f"{utente['name']} {utente['cognome']}" if utente else 'Un dipendente'
    etichetta_tipo = 'ferie' if richiesta['tipo'] == 'ferie' else 'permesso'
    periodo = richiesta['data_inizio'] if richiesta['data_inizio'] == richiesta['data_fine'] else f"dal {richiesta['data_inizio']} al {richiesta['data_fine']}"

    crea_notifica(
        f"{nome_completo} chiede di rimuovere {etichetta_tipo} già approvato/a ({periodo})",
        tipo='rimozione',
        user_id=user_id,
        data_riferimento=richiesta['data_inizio']
    )

    return redirect(url_for('richieste_page'))


@app.route('/richieste/eventi')
def richieste_eventi():
    """Restituisce, in JSON, ferie e permessi (approvati e in attesa) del
    dipendente collegato, da mostrare nel calendario dei giorni liberi."""
    if 'user' not in session or session.get('tipo_utente') == 'admin':
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    user_id = session['user'].get('id')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        '''SELECT tipo, data_inizio, data_fine, stato, ora_inizio, ora_fine, motivo
           FROM richieste
           WHERE user_id = ? AND stato IN ('approvata', 'in_attesa')''',
        (user_id,)
    )
    righe = cursor.fetchall()
    conn.close()

    return jsonify({
        'ok': True,
        'eventi': [
            {
                'tipo': r['tipo'],
                'data_inizio': r['data_inizio'],
                'data_fine': r['data_fine'],
                'stato': r['stato'],
                'ora_inizio': r['ora_inizio'],
                'ora_fine': r['ora_fine'],
                'motivo': r['motivo'],
            }
            for r in righe
        ]
    })


@app.route('/admin/calendario/eventi')
def admin_calendario_eventi():
    """Restituisce, in JSON, tutte le ferie e i permessi approvati di tutti
    i dipendenti, da mostrare nel calendario aziendale del pannello admin."""
    if session.get('tipo_utente') != 'admin':
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        '''SELECT richieste.tipo, richieste.data_inizio, richieste.data_fine,
                  richieste.ora_inizio, richieste.ora_fine, richieste.motivo,
                  users.name, users.cognome
           FROM richieste JOIN users ON users.id = richieste.user_id
           WHERE richieste.stato = 'approvata'
           ORDER BY richieste.data_inizio ASC'''
    )
    righe = cursor.fetchall()
    conn.close()

    return jsonify({
        'ok': True,
        'eventi': [
            {
                'tipo': r['tipo'],
                'data_inizio': r['data_inizio'],
                'data_fine': r['data_fine'],
                'ora_inizio': r['ora_inizio'],
                'ora_fine': r['ora_fine'],
                'motivo': r['motivo'],
                'nome': r['name'],
                'cognome': r['cognome'],
            }
            for r in righe
        ]
    })


@app.route('/admin/notifiche/segna_lette', methods=['POST'])
def segna_notifiche_lette():
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE notifiche SET letta = 1 WHERE letta = 0')
    conn.commit()
    conn.close()

    return redirect(url_for('admin_panel'))


@app.route('/admin/notifiche/elimina/<int:notifica_id>', methods=['POST'])
def elimina_notifica(notifica_id):
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM notifiche WHERE id = ?', (notifica_id,))
    conn.commit()
    eliminata = cursor.rowcount > 0
    conn.close()

    if not eliminata:
        return jsonify({'ok': False, 'errore': 'Notifica non trovata'}), 404

    return redirect(url_for('admin_panel'))


@app.route('/admin/notifiche/elimina_tutte', methods=['POST'])
def elimina_tutte_notifiche():
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM notifiche')
    conn.commit()
    conn.close()

    return redirect(url_for('admin_panel'))


@app.route('/admin/notifiche/<int:notifica_id>/segnala', methods=['POST'])
def segnala_notifica(notifica_id):
    """Invia un'email al dipendente coinvolto nella notifica, come segnalazione formale.
    Permesso solo per ritardi ('alert'), assenze ('assenza') o straordinari ('straordinario')."""
    redirect_response = require_admin()
    if redirect_response:
        return redirect_response

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM notifiche WHERE id = ?', (notifica_id,))
    notifica = cursor.fetchone()

    if notifica is None:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Notifica non trovata'}), 404

    if notifica['tipo'] not in ('alert', 'assenza', 'straordinario', 'monte_ore'):
        conn.close()
        return jsonify({
            'ok': False,
            'errore': 'Le segnalazioni possono essere fatte solo per ritardi, assenze, straordinari o monte ore non rispettato'
        }), 400

    if not notifica['user_id']:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Nessun dipendente associato a questa notifica'}), 400

    cursor.execute('SELECT * FROM users WHERE id = ?', (notifica['user_id'],))
    user = cursor.fetchone()

    if user is None:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Dipendente non trovato'}), 404

    cursor.execute('UPDATE notifiche SET segnalata = 1 WHERE id = ?', (notifica_id,))
    conn.commit()
    conn.close()

    testo_messaggio = f'Segnalazione automatica: "{notifica["messaggio"]}". Se pensi si tratti di un errore, rispondi pure qui.'
    crea_messaggio(notifica['user_id'], 'admin', testo_messaggio)

    return redirect(url_for('admin_panel'))


@app.route('/admin/utente/<int:user_id>/posizione')
def admin_posizione_utente(user_id):
    """Restituisce la posizione attuale del dipendente, ma solo se in questo
    momento risulta al lavoro (tra il timbro di entrata e quello di uscita
    della giornata odierna). Fuori da questa finestra non viene restituita
    alcuna posizione, anche se in passato ne sono state registrate."""
    if session.get('tipo_utente') != 'admin':
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT name, cognome FROM users WHERE id = ?', (user_id,))
    utente = cursor.fetchone()
    if utente is None:
        conn.close()
        return jsonify({'ok': False, 'errore': 'Utente non trovato'}), 404

    now = datetime.now()
    oggi_iso = now.date().isoformat()

    cursor.execute(
        '''SELECT tipo, creato_il, lat, lng FROM timbrature
           WHERE user_id = ? AND creato_il LIKE ? ORDER BY creato_il ASC''',
        (user_id, f'{oggi_iso}%')
    )
    punches_oggi = [dict(r) for r in cursor.fetchall()]
    conn.close()

    stato, _ = calcola_stato_giornata(punches_oggi, now)
    if stato not in ('in_corso', 'in_pausa'):
        return jsonify({
            'ok': False,
            'errore': 'Il dipendente non è al lavoro in questo momento: nessuna posizione da mostrare.'
        }), 404

    punches_con_posizione = [p for p in punches_oggi if p['lat'] is not None and p['lng'] is not None]
    if not punches_con_posizione:
        return jsonify({
            'ok': False,
            'errore': 'Il dipendente è al lavoro, ma nessuna delle timbrature odierne include una posizione.'
        }), 404

    ultima = punches_con_posizione[-1]

    etichette = {
        'entrata': "Entrata",
        'pausa_inizio': "Inizio pausa",
        'pausa_fine': "Fine pausa",
        'uscita': "Uscita",
    }
    creato_il = datetime.fromisoformat(ultima['creato_il'])

    return jsonify({
        'ok': True,
        'lat': ultima['lat'],
        'lng': ultima['lng'],
        'tipo': etichette.get(ultima['tipo'], ultima['tipo']),
        'data': creato_il.strftime('%d/%m/%Y'),
        'ora': creato_il.strftime('%H:%M'),
        'nome_completo': f"{utente['name']} {utente['cognome']}"
    })


@app.route('/admin/chat/<int:user_id>/invia', methods=['POST'])
def admin_chat_invia(user_id):
    if session.get('tipo_utente') != 'admin':
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    testo = request.form.get('testo', '').strip()
    if not testo:
        return jsonify({'ok': False, 'errore': 'Il messaggio non può essere vuoto'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT 1 FROM users WHERE id = ?', (user_id,))
    esiste = cursor.fetchone()
    conn.close()
    if not esiste:
        return jsonify({'ok': False, 'errore': 'Utente non trovato'}), 404

    crea_messaggio(user_id, 'admin', testo)
    return jsonify({'ok': True})


@app.route('/admin/chat/<int:user_id>/nuovi')
def admin_chat_nuovi(user_id):
    if session.get('tipo_utente') != 'admin':
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE messaggi SET letto_admin = 1 WHERE user_id = ? AND mittente = 'utente'",
        (user_id,)
    )
    conn.commit()
    cursor.execute(
        'SELECT mittente, testo, creato_il FROM messaggi WHERE user_id = ? ORDER BY creato_il ASC',
        (user_id,)
    )
    righe = cursor.fetchall()
    conn.close()

    return jsonify({
        'ok': True,
        'messaggi': [
            {'mittente': r['mittente'], 'testo': r['testo'], 'ora': r['creato_il'][11:16]}
            for r in righe
        ]
    })


@app.route('/chat/invia', methods=['POST'])
def chat_invia():
    if 'user' not in session or session.get('tipo_utente') == 'admin':
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    testo = request.form.get('testo', '').strip()
    if not testo:
        return jsonify({'ok': False, 'errore': 'Il messaggio non può essere vuoto'}), 400

    user_id = session['user'].get('id')
    crea_messaggio(user_id, 'utente', testo)
    return jsonify({'ok': True})


@app.route('/chat/nuovi')
def chat_nuovi():
    if 'user' not in session or session.get('tipo_utente') == 'admin':
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    user_id = session['user'].get('id')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE messaggi SET letto_utente = 1 WHERE user_id = ? AND mittente = 'admin'",
        (user_id,)
    )
    conn.commit()
    cursor.execute(
        'SELECT mittente, testo, creato_il FROM messaggi WHERE user_id = ? ORDER BY creato_il ASC',
        (user_id,)
    )
    righe = cursor.fetchall()
    conn.close()

    return jsonify({
        'ok': True,
        'messaggi': [
            {'mittente': r['mittente'], 'testo': r['testo'], 'ora': r['creato_il'][11:16]}
            for r in righe
        ]
    })


@app.route('/chat/non_lette')
def chat_non_lette():
    if 'user' not in session or session.get('tipo_utente') == 'admin':
        return jsonify({'ok': True, 'non_lette': 0})

    user_id = session['user'].get('id')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT COUNT(*) AS n FROM messaggi WHERE user_id = ? AND mittente = 'admin' AND letto_utente = 0",
        (user_id,)
    )
    n = cursor.fetchone()['n']
    conn.close()

    return jsonify({'ok': True, 'non_lette': n})


@app.route('/login', methods=['POST'])
def login():
    email = request.form.get('email', '').strip().lower()
    password = request.form.get('password', '').strip()

    if not email or not password:
        return jsonify({'ok': False, 'errore': 'Email e password sono obbligatorie'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE email = ?', (email,))
    user = cursor.fetchone()
    conn.close()

    if user is None or not check_password_hash(user['password'], password):
        return jsonify({'ok': False, 'errore': 'Credenziali non valide'}), 401

    session.permanent = True
    session['user'] = {
        'id': user['id'],
        'email': user['email'],
        'name': user['name'] if 'name' in user.keys() else user['email'],
        'authenticated_at': datetime.utcnow().isoformat()
    }
    session['tipo_utente'] = 'utente'

    return redirect(url_for('timbro_page'))


@app.route('/timbro')
def timbro_page():
    if 'user' not in session:
        return redirect(url_for('login_page'))
    if session.get('tipo_utente') == 'admin':
        return redirect(url_for('admin_panel'))

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT privacy_accettata FROM users WHERE id = ?', (session['user']['id'],))
    riga = cursor.fetchone()
    conn.close()
    privacy_da_accettare = riga is None or not riga['privacy_accettata']

    return render_template('timbro.html', is_admin=False, privacy_da_accettare=privacy_da_accettare)


@app.route('/account/privacy/accetta', methods=['POST'])
def accetta_privacy():
    if 'user' not in session:
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE users SET privacy_accettata = 1 WHERE id = ?', (session['user']['id'],))
    conn.commit()
    conn.close()

    return jsonify({'ok': True})


@app.route('/timbro/stato')
def timbro_stato():
    if 'user' not in session:
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401
    if session.get('tipo_utente') == 'admin':
        return jsonify({'ok': False, 'errore': 'Non disponibile per l\'amministratore'}), 403

    user_id = session['user'].get('id')
    now = datetime.now()
    oggi_iso = now.date().isoformat()

    if ha_ferie_approvata(user_id, oggi_iso):
        return jsonify({
            'ok': True,
            'is_admin': False,
            'stato': 'in_ferie',
            'secondi_lavorati': 0,
            'timbrature_oggi': []
        })

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT tipo, creato_il FROM timbrature WHERE user_id = ? AND creato_il LIKE ? ORDER BY creato_il ASC",
        (user_id, f'{oggi_iso}%')
    )
    righe = cursor.fetchall()
    conn.close()

    punches = [dict(r) for r in righe]
    stato, secondi_lavorati = calcola_stato_giornata(punches, now)

    return jsonify({
        'ok': True,
        'is_admin': False,
        'stato': stato,
        'secondi_lavorati': secondi_lavorati,
        'timbrature_oggi': [
            {'tipo': p['tipo'], 'ora': datetime.fromisoformat(p['creato_il']).strftime('%H:%M')}
            for p in punches
        ]
    })


@app.route('/timbro/storico')
def timbro_storico():
    if 'user' not in session:
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    if session.get('tipo_utente') == 'admin':
        return jsonify({'ok': False, 'errore': "Non disponibile per l'amministratore"}), 403

    user_id = session['user'].get('id')
    now = datetime.now()
    da_data = (now.date() - timedelta(days=13)).isoformat()

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT tipo, creato_il FROM timbrature WHERE user_id = ? AND creato_il >= ? ORDER BY creato_il ASC",
        (user_id, da_data)
    )       
    righe = cursor.fetchall()
    conn.close()

    per_giorno = {}
    for r in righe:
        giorno = r['creato_il'][:10]
        per_giorno.setdefault(giorno, []).append(dict(r))

    giorni = []
    for giorno, punches in per_giorno.items():
        stato, secondi = calcola_stato_giornata(punches, now)
        giorni.append({
            'data': giorno,
            'stato': stato,
            'secondi_lavorati': secondi,
            'incompleto': stato in ('in_corso', 'in_pausa'),
            'timbrature': [
                {'tipo': p['tipo'], 'ora': datetime.fromisoformat(p['creato_il']).strftime('%H:%M')}
                for p in punches
            ]
        })

    giorni.sort(key=lambda g: g['data'], reverse=True)

    return jsonify({'ok': True, 'giorni': giorni})


@app.route('/timbro/azione', methods=['POST'])
def timbro_azione():
    if 'user' not in session:
        return jsonify({'ok': False, 'errore': 'Non autenticato'}), 401

    tipo = request.form.get('tipo', '').strip()
    if tipo not in ('entrata', 'pausa_inizio', 'pausa_fine', 'uscita'):
        return jsonify({'ok': False, 'errore': 'Tipo di timbratura non valido'}), 400
    if session.get('tipo_utente') == 'admin':
        return jsonify({'ok': False, 'errore': "Non disponibile per l'amministratore"}), 403

    lat = None
    lng = None
    lat_raw = request.form.get('lat', '').strip()
    lng_raw = request.form.get('lng', '').strip()
    if lat_raw and lng_raw:
        try:
            lat_val = float(lat_raw)
            lng_val = float(lng_raw)
            if -90 <= lat_val <= 90 and -180 <= lng_val <= 180:
                lat, lng = lat_val, lng_val
        except ValueError:
            pass

    user_id = session['user'].get('id')
    adesso = datetime.now()
    oggi_iso = adesso.date().isoformat()

    if ha_ferie_approvata(user_id, oggi_iso):
        return jsonify({
            'ok': False,
            'errore': 'Oggi risulti in ferie: non puoi timbrare.',
            'stato': 'in_ferie'
        }), 403

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT tipo, creato_il FROM timbrature WHERE user_id = ? AND creato_il LIKE ? ORDER BY creato_il ASC",
        (user_id, f'{oggi_iso}%')
    )
    punches_oggi = [dict(r) for r in cursor.fetchall()]
    stato_attuale, _ = calcola_stato_giornata(punches_oggi, adesso)

    if tipo not in TRANSIZIONI_AMMESSE[stato_attuale]:
        conn.close()
        messaggi_stato = {
            'non_iniziato': "Devi prima timbrare l'entrata",
            'in_corso': "Sei già al lavoro: puoi mettere in pausa o timbrare l'uscita",
            'in_pausa': "Sei in pausa: puoi solo terminare la pausa",
            'terminato': "Hai già timbrato l'uscita: la giornata è conclusa, non puoi più timbrare fino a domani",
        }
        return jsonify({
            'ok': False,
            'errore': messaggi_stato.get(stato_attuale, 'Azione non consentita in questo momento'),
            'stato': stato_attuale
        }), 409

    cursor.execute(
        'INSERT INTO timbrature (user_id, tipo, creato_il, lat, lng) VALUES (?, ?, ?, ?, ?)',
        (user_id, tipo, adesso.isoformat(), lat, lng)
    )
    conn.commit()

    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()
    conn.close()

    nome_completo = f"{user['name']} {user['cognome']}" if user else 'Un dipendente'
    etichette = {
        'entrata': "l'entrata",
        'pausa_inizio': "l'inizio pausa",
        'pausa_fine': "la fine pausa",
        'uscita': "l'uscita",
    }
    ora_str = adesso.strftime('%H:%M')
    messaggio = f"{nome_completo} ha timbrato {etichette[tipo]} alle {ora_str}"
    tipo_notifica = 'info'

    orario_giorno = get_orario_giorno(user, oggi_iso) if user else None
    permesso_approvato_oggi = ha_permesso_approvato(user_id, oggi_iso)

    if tipo == 'entrata' and orario_giorno and orario_giorno['orario_inizio']:
        try:
            atteso = datetime.strptime(orario_giorno['orario_inizio'], '%H:%M').time()
            atteso_dt = adesso.replace(hour=atteso.hour, minute=atteso.minute, second=0, microsecond=0)
            ritardo_minuti = (adesso - atteso_dt).total_seconds() / 60
            if ritardo_minuti > 10:
                messaggio += f" (in ritardo di {int(ritardo_minuti)} minuti)"
                tipo_notifica = 'alert'
        except ValueError:
            pass

    elif tipo == 'uscita':
        punches_oggi_con_uscita = punches_oggi + [{'tipo': tipo, 'creato_il': adesso.isoformat()}]
        _, secondi_lavorati_oggi = calcola_stato_giornata(punches_oggi_con_uscita, adesso)
        ore_lavorate = secondi_lavorati_oggi // 3600
        minuti_lavorati = (secondi_lavorati_oggi % 3600) // 60
        messaggio += f" ({ore_lavorate}h {minuti_lavorati}m lavorate oggi)"

        if orario_giorno and orario_giorno['orario_inizio'] and orario_giorno['orario_fine']:
            try:
                inizio_previsto = datetime.strptime(orario_giorno['orario_inizio'], '%H:%M')
                fine_prevista = datetime.strptime(orario_giorno['orario_fine'], '%H:%M')
                monte_ore_previsto_minuti = (fine_prevista - inizio_previsto).total_seconds() / 60

                if orario_giorno['orario_pausa_inizio'] and orario_giorno['orario_pausa_fine']:
                    pausa_inizio_prevista = datetime.strptime(orario_giorno['orario_pausa_inizio'], '%H:%M')
                    pausa_fine_prevista = datetime.strptime(orario_giorno['orario_pausa_fine'], '%H:%M')
                    monte_ore_previsto_minuti -= (pausa_fine_prevista - pausa_inizio_prevista).total_seconds() / 60

                if orario_giorno.get('permesso_inizio') and orario_giorno.get('permesso_fine'):
                    permesso_inizio_previsto = datetime.strptime(orario_giorno['permesso_inizio'], '%H:%M')
                    permesso_fine_previsto = datetime.strptime(orario_giorno['permesso_fine'], '%H:%M')
                    monte_ore_previsto_minuti -= (permesso_fine_previsto - permesso_inizio_previsto).total_seconds() / 60

                minuti_lavorati_oggi = secondi_lavorati_oggi / 60
                extra_minuti = minuti_lavorati_oggi - monte_ore_previsto_minuti

                if extra_minuti > 15:
                    ore_extra, minuti_extra = divmod(int(extra_minuti), 60)
                    messaggio += f" — straordinario di {ore_extra}h {minuti_extra}m sul monte ore previsto"
                    tipo_notifica = 'straordinario'
                elif extra_minuti < -15 and not permesso_approvato_oggi:
                    ore_mancanti, minuti_mancanti = divmod(int(abs(extra_minuti)), 60)
                    messaggio += f" — monte ore non rispettato: mancano {ore_mancanti}h {minuti_mancanti}m rispetto all'orario previsto"
                    tipo_notifica = 'monte_ore'
            except ValueError:
                pass

    crea_notifica(messaggio, tipo=tipo_notifica, user_id=user_id, data_riferimento=adesso.date().isoformat(), lat=lat, lng=lng)

    punches_oggi.append({'tipo': tipo, 'creato_il': adesso.isoformat()})
    nuovo_stato, secondi_lavorati = calcola_stato_giornata(punches_oggi, adesso)

    return jsonify({
        'ok': True,
        'registrata': True,
        'stato': nuovo_stato,
        'secondi_lavorati': secondi_lavorati,
        'ora': ora_str
    })


@app.route('/account/password', methods=['GET'])
def cambia_password_page():
    if 'user' not in session:
        return redirect(url_for('login_page'))
    if session.get('tipo_utente') == 'admin':
        return redirect(url_for('admin_panel'))
    return render_template('cambia_password.html')


@app.route('/account/password', methods=['POST'])
def cambia_password():
    if 'user' not in session:
        return redirect(url_for('login_page'))
    if session.get('tipo_utente') == 'admin':
        return redirect(url_for('admin_panel'))

    password_attuale = request.form.get('password_attuale', '')
    nuova_password = request.form.get('nuova_password', '')
    conferma_password = request.form.get('conferma_password', '')
    user_id = session['user'].get('id')

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    user = cursor.fetchone()

    errore = None
    if user is None:
        errore = 'Utente non trovato'
    elif not check_password_hash(user['password'], password_attuale):
        errore = 'La password attuale non è corretta'
    elif len(nuova_password) < 8:
        errore = 'La nuova password deve avere almeno 8 caratteri'
    elif nuova_password != conferma_password:
        errore = 'Le due password inserite non coincidono'
    elif check_password_hash(user['password'], nuova_password):
        errore = 'La nuova password deve essere diversa da quella attuale'

    if errore:
        conn.close()
        return render_template('cambia_password.html', errore=errore)

    cursor.execute(
        'UPDATE users SET password = ? WHERE id = ?',
        (generate_password_hash(nuova_password), user_id)
    )
    conn.commit()
    conn.close()

    return render_template('cambia_password.html', successo='Password aggiornata con successo.')


@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return redirect(url_for('index'))


if __name__ == '__main__':
    init_db()
    app.run('127.0.0.1', 5000, debug=True)